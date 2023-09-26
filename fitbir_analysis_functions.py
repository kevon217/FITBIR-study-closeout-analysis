# -*- coding: utf-8 -*-
"""
Created on Thu Mar 18 10:46:20 2021

@author: kevarmen1
"""

## THIS FILE HOLDS FUNCTIONS USED FOR LOADING/PROCESSING FITBIR DATA AS CSV ## 
import pandas as pd
import numpy as np
import os
import shutil
from loguru import logger
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import re
sep = "," # for separating log messages with comma for opening as csv and filtering


####### FOLDER/DIRECTORY FUNCTIONS #######

# will create new folders everytime analysis is run
def create_folder(folder_path):
    adjusted_folder_path = folder_path
    folder_found = os.path.isdir(adjusted_folder_path)
    counter = 0
    while folder_found == True:
        counter = counter + 1
        adjusted_folder_path = folder_path + ' (' + str(counter) + ')'
        folder_found = os.path.isdir(adjusted_folder_path)
    os.mkdir(adjusted_folder_path)
    return adjusted_folder_path

# will copy directories for final closeout package
def copyDirectory(src, dest):
    try:
        shutil.copytree(src, dest)
    # Directories are the same
    except shutil.Error as e:
        logger.error('Directory not copied. Error: %s' % e)
    # Any error saying that the directory doesn't exist
    except OSError as e:
        logger.error('Directory not copied. Error: %s' % e)

####### CSV OPERATIONS #######

# filter query tool csv output with csv of dataset list
def id_filter(df, filename, filter_ID_list):    
      id_idx = df.loc[:,'Dataset'].isin(filter_ID_list)
      df = df[id_idx].reset_index(drop = True) # need to reset index when filtering for downstream operations e.g., highlighting row/col coordinates/indices
      logger.info(sep.join(['Filtering by Dataset IDs:', filename]))
      return df
    # figure out if there are any issues with resetting indices or not

# (1) load data (2) filter? (3) count DEs/records/unique guids (4) figure out repeatable and non-repeatable group columns
def loadFile(filename,n_skiprows,filter_ID_1_0,filter_ID_list,):
    if filter_ID_1_0:
        df = pd.read_csv(filename,low_memory=False,skiprows = n_skiprows, dtype = 'object')
        df = id_filter(df, filename, filter_ID_list)
    else:
        df = pd.read_csv(filename,low_memory=False,skiprows = n_skiprows, dtype = 'object')
    col_guid = [col for col in df.columns if '.GUID' in col.upper()][0] # col.upper() used because some data repo files can have GUID with different case e.g., Main.Guid 
    try: 
        col_assocguid = [col for col in df.columns if '.ASSOCIATED GUID' in col.upper()][0] # assoc guid only in QT output
    except: 
        col_assocguid = [] # no assoc guid col in Data Repo files
    re_subjID = re.compile("Subject[_]?ID", re.IGNORECASE) # TODO: use same logic for col
    col_subjID =  list(filter(re_subjID.search, df.columns))[0]
    n_records = len(df)
    # get repeatable/non-repeatable group cols
    cols_rg = [col for col in df.columns if re.search(r'(_\d*$)',col)]
    cols_nrg = list(np.setdiff1d(df.columns,cols_rg))
    cols_rg = pd.Series([re.split(r'(_\d*$)',col)[0] for col in cols_rg]).drop_duplicates().tolist() # remove "_\d" extension from repeatable groups
    return(df, cols_rg, cols_nrg, col_guid, col_assocguid, col_subjID, n_records)

# check if dataframe is flat e.g., TrialNum_1, TrialNum_2, vs. unflat e.g., TrialNum
def checkFlat(df):
    cols_rg = [col for col in df.columns if re.search(r'(_\d*$)',col)] # find repeatable group columns
    if len(cols_rg) == 0:
        isflat = 0
    else:
        isflat = 1
    return isflat

# convert flattened data to unflattend; adds columns for precise filtering: Record/Repeatable Group/Validation Row
def flat2unflat(df_flat, filename, cols_rg, cols_nrg, col_guid, col_subjID, col_assocguid):
    df_temp = df_flat.copy() 
    isflat = checkFlat(df_temp)
    if 'record' not in df_temp.columns:  # DR files already have 'record' column
        df_temp.insert(0,'record',range(1,len(df_temp)+1)) # keep track of records
    else: 
        df_temp['record'] = range(1,len(df_temp)+1) # replace DR files record label with numbers
    # cols_df = df_temp.columns # keep track of col order
    # cols_unflat_order = list(dict.fromkeys([re.sub(r'(_\d*$)', '', col) for col in cols_df])) # unflattened order of columns
    cols_order = list(dict.fromkeys([re.sub(r'(_\d*$)', '', col) for col in df_temp.columns]))
    cols_order.remove('record')
    cols_reorder = ['record','Repeatable Group','Validation Row'] + cols_order # reordered columns
    if isflat:
        logger.info(sep.join(["Repeatable groups present in: ", filename]))
        df_unflat = pd.wide_to_long(df_temp, stubnames = cols_rg, i = "record", 
                                    j="Repeatable Group", suffix = '\d+', sep = "_")
        df_unflat = df_unflat.sort_values(by=['record','Repeatable Group']).reset_index()
        
        # remove empty repeatable group rows created by pd.wide_to_long; it will generate the max number of repeatable groups in dataset for each record; WARNING: this would remove intentionally blank repeatable group rows.
        rows_rg2plus_idx = df_unflat[df_unflat['Repeatable Group'] != 1].index.tolist()
        rows_rg2plus_nonempty_idx = df_unflat.loc[rows_rg2plus_idx].dropna(subset = cols_rg, how = 'all').index.tolist()
        rows_drop =  list(set(rows_rg2plus_idx) - set(rows_rg2plus_nonempty_idx))
        df_unflat.drop(labels = rows_drop, axis = 0, inplace = True)
       
        df_unflat['Validation Row'] = range(3,len(df_unflat)+3) # this corresponds to row number in actual submission; used to correlate val results with 'All Submitted Data' position for highlighting
        df_unflat = df_unflat[cols_reorder] # original DE order
        if len(col_assocguid) != 0:
            cols_rm_ffill = set(cols_nrg).difference(set(['Study ID', 'Dataset', col_guid, col_subjID, col_assocguid]))
        else:
            try:
                cols_rm_ffill = set(cols_nrg).difference(set(['Study ID', 'Dataset', col_guid, col_subjID])) # no associated guid
            except: 
                cols_rm_ffill = set(cols_nrg).difference(set(['Study ID', 'Dataset', col_guid])) # no subject ID or associated guid
        df_unflat.loc[df_unflat['Repeatable Group'] > 1, cols_rm_ffill] = np.nan # remove ffill from rows for non-repeatable group DEs EXCEPT Study ID/Dataset/GUID/ASSOCIATED GUID/SUBJECT ID
        df_unflat.reset_index(inplace = True, drop = True)
    else:
        cols_order = list(df_temp.columns) # keep track of col order
        logger.info(sep.join(["No repeatable groups present in: ", filename]))
        df_unflat = df_temp # if original is not flat, then it is already unflattened
        df_unflat['Repeatable Group'] = 1
        df_unflat['Validation Row'] = range(3,len(df_unflat)+3) # this corresponds to row number in actual submission; used to correlate val results with 'All Submitted Data' position for highlighting
        df_unflat = df_unflat[cols_reorder] # original DE order
    return df_unflat

# convert unflattened query tool file to validation ready csv format
def unflat_reformat(df, formName):
    df_temp = df.iloc[:,1:].copy() # removes 'Study ID' col
    try:
        col_assocguid = [col for col in df.columns if 'ASSOCIATED GUID' in col.upper()][0] # can't submit to FITBIR with ASSOCIATED.GUID column
        df_temp.drop(columns=[col_assocguid], inplace = True)
        df_val_csv = df_temp
    except:
        df_val_csv = df_temp
        logger.info(sep.join(['No ASSOCIATED GUID Column', formName]))
    
    # get variable headers without form structure name
    name_split = lambda x: x.rsplit('.')[-2:] # get's words surrounded by last period
    name_join = lambda x: '.'.join(x) # joins these two words e.g., Main.GUID
    var_names = df_temp.columns   
    var_name_split = list(map(name_split, var_names))
    var_name_join = list(map(name_join, var_name_split))
    df_val_csv.columns = var_name_join
    
    # create empty col and row 
    empty_row = pd.DataFrame(np.empty([2,df_val_csv.shape[1]]),columns = var_name_join)
    empty_row[:] = np.nan
    
    # modiy query format to look like format for DR and validation
    df_val_csv = pd.concat([empty_row,df_val_csv],axis = 0, ignore_index = True)
    df_val_csv.loc[0,'Dataset'] = formName
    df_val_csv.loc[1,'Dataset'] = 'record'
    df_val_csv.iloc[1,1:] = var_name_join[1:]
    
    return(df_val_csv)

####### CLOSEOUT ANALYSIS FUNCTIONS #######

# find empty columns and remove
def currentDataSubmitted(df_unflat, col_guid): 
    df_temp = df_unflat.copy()
    dfReduced = df_temp.dropna(axis=1, how ="all")
    emptyColumns = []
    for col in df_temp.columns:
        if (df_temp[col].isnull().all()):
            emptyColumns.append(col)
    if {'Study ID','Dataset','record','Repeatable Group','Validation Row'}.issubset(df_unflat):    
        n_DE_total = len(df_unflat.drop(columns=['Study ID','Dataset','record','Repeatable Group','Validation Row']).columns)
        if dfReduced.empty:
            n_DE_data = 0
        else:
            n_DE_data = len(dfReduced.drop(columns=['Study ID','Dataset','record','Repeatable Group','Validation Row']).columns)
    else:
        n_DE_total = len(df_unflat.columns)
        n_DE_data = len(dfReduced.columns)
    n_GUIDs_uniq = len((df_unflat.loc[:,col_guid].unique()))    
    return(dfReduced,emptyColumns,n_DE_total,n_DE_data,n_GUIDs_uniq)

# find duplicates
def returnDuplicated(df_unflat, col_guid):
    df_temp = df_unflat.copy()
    df_temp[col_guid] = df_temp[col_guid].ffill() # ffill GUID for repeatable groups
    col_subset = list(df_temp.columns.difference(['record','Repeatable Group','Validation Row','Study ID','Dataset']))
    dup_idx = df_temp.duplicated(keep = False, subset = col_subset)
    dfDuplicated = df_temp[dup_idx].sort_values(by=['Validation Row', 'Repeatable Group'])
    if ('Study ID' in dfDuplicated.columns) & ('Dataset' in dfDuplicated.columns):
            dfDuplicated['Study ID'] = dfDuplicated['Study ID'].ffill()
            dfDuplicated['Dataset'] = dfDuplicated['Dataset'].ffill()
    if (len(dfDuplicated)> 0):
        n_dup_rows = len(dfDuplicated['record'].unique())
    else:
        n_dup_rows = 0
    dup_rows = dfDuplicated.index
    return(dfDuplicated, dup_rows, n_dup_rows)

# find partial data (leave in flattened format)
def GUIDswMissingDatainDEs(df_flat, cols_rg, cols_nrg):
    df_temp = df_flat.copy()
    dfReduced = df_temp.dropna(axis=1, how ="all") # remove empty columns
    n_nulls_col = dfReduced.isnull().sum().sort_values(ascending=False)# number of nulls in cols sorted by decreasing order
    dfReduced = dfReduced.reindex(n_nulls_col.index, axis=1) # sort df by # of partial DEs decreasing order
    col_guid = [col for col in dfReduced.columns if '.GUID' in col.upper()][0]
    # col_assocguid = [col for col in cols_flat if '.ASSOCIATED GUID' in col.upper()][0]
    
    # this finds missing data in data elements
    dfReduced_nulls = dfReduced.isnull() # true is value is null
    rmv_col = np.where(np.array(n_nulls_col == 0))
    dfReduced_partial_DE = dfReduced_nulls.drop(dfReduced_nulls.columns[rmv_col],axis=1) # only cols with partial DEs
    if dfReduced_partial_DE.shape[1] > 0:
        guids = df_flat[col_guid].iloc[:]
        dfReduced_partial_DE =  dfReduced_partial_DE.apply(lambda x: x*guids)
        missingGuids =  dfReduced_partial_DE.apply(lambda x: pd.Series(x.unique()).sort_values().reset_index(drop=True))
        missingGuids =  missingGuids.iloc[1:,:]
    else:
        missingGuids = pd.DataFrame()
    cols_partial_DE = dfReduced_partial_DE.columns # for highlighting partial DE in "All Submitted Data" in output excel file
    cols_partial_DE = pd.Series([re.split(r'(_\d*$)',col)[0] for col in cols_partial_DE]).drop_duplicates().tolist() # remove "_\d" from rg cols for highlighting single column in unflattened presentation
    guidCounts = df_flat[col_guid].value_counts()
    return(missingGuids, cols_partial_DE, guidCounts)

# find GUIDs in basefile missing in comparison file   
def GUIDsMissingFromCompFile(fn_base_file, filename, df_base_file, df_flat):
    BaseFile = df_base_file
    ComparisonFile = df_flat
    col_guid_BaseFile = [col for col in BaseFile.columns if '.GUID' in col.upper()][0]
    col_guid_ComparisonFile = [col for col in ComparisonFile.columns if '.GUID' in col.upper()][0]
    df_GUIDs_miss_compfile = BaseFile[~BaseFile.loc[:,col_guid_BaseFile].isin(ComparisonFile.loc[:,col_guid_ComparisonFile])]
    n_GUIDs_miss_compfile = len( df_GUIDs_miss_compfile.iloc[:,2].value_counts())
    return(df_GUIDs_miss_compfile, n_GUIDs_miss_compfile)
  
# find GUIDs in comparison file missing from basefile  
def GUIDsMissingFromBase(fn_base_file, filename, df_base_file, df_flat):
    BaseFile = df_base_file
    ComparisonFile = df_flat
    col_guid_BaseFile = [col for col in BaseFile.columns if '.GUID' in col.upper()][0]
    col_guid_ComparisonFile = [col for col in ComparisonFile.columns if '.GUID' in col.upper()][0]
    df_GUIDs_miss_basefile = ComparisonFile[~ComparisonFile.loc[:,col_guid_ComparisonFile].isin(BaseFile.loc[:,col_guid_BaseFile])]
    n_GUIDs_miss_basefile = len(df_GUIDs_miss_basefile.iloc[:,2].value_counts())
    return(df_GUIDs_miss_basefile, n_GUIDs_miss_basefile)

# check GUID subjID correspondence
def guid_subjID_check(df_flat, qt_dr_1_0, col_guid, col_assocguid, col_subjID):
    df_temp = df_flat.copy()
    if {'Form Structure'}.issubset(df_temp): # for df_GUIDSubjID_master df
        df_temp['record'] = df_temp.index + 1 # indices correspond to record number in each form separately. +1 to offset 0 index start
        if len(col_assocguid) == 0: 
            try: 
                df_temp = df_temp[['Study ID','Form Structure','Dataset', 'record', col_guid, col_subjID]] # for QT files
            except: 
                df_temp = df_temp[['Form Structure','record',col_guid,col_subjID]] # for DR files
        else:
            try: 
                df_temp = df_temp[['Study ID','Form Structure','Dataset', 'record', col_guid, col_assocguid, col_subjID]] # for QT files
            except: 
                df_temp = df_temp[['Form Structure','record',col_guid,col_subjID]] # for DR files
    else: # for within single Form Structure test
        df_temp['record'] = range(1,len(df_temp)+1) # keep track of records
        if len(col_assocguid) == 0: 
            if qt_dr_1_0: 
                df_temp = df_temp[['Study ID','Dataset', 'record', col_guid, col_subjID]] # for QT files
            else: 
                df_temp = df_temp[['record',col_guid,col_subjID]] # for DR files
        else:
            if qt_dr_1_0:
                df_temp = df_temp[['Study ID','Dataset', 'record', col_guid, col_assocguid, col_subjID]] # for QT files
            else: 
                df_temp = df_temp[['record',col_guid,col_subjID]] # for DR files
     
   
    if (len(col_assocguid) == 0) | (col_assocguid == 'ASSOCIATED GUID'):
       # find SubjIDs with multiple GUIDs
       cnt_subjID =  df_temp.groupby(col_subjID)[col_guid].unique().apply(len)
       list_subjID = cnt_subjID.index[cnt_subjID>1].tolist()
       df_subjID_guid = pd.DataFrame(columns = df_temp.columns)
       for i in range(0,len(list_subjID)):
           subjID = list_subjID[i]
           temp = df_temp[df_temp[col_subjID]==subjID]
           temp = temp[-temp.duplicated()]
           df_subjID_guid = pd.concat([df_subjID_guid, temp],ignore_index = False)
           
       # find GUIDs with multiple SubjIDs  
       cnt_GUID = df_temp.groupby(col_guid)[col_subjID].unique().apply(len)
       list_GUID = cnt_GUID.index[cnt_GUID>1].tolist()
       df_guid_subjID = pd.DataFrame(columns = df_temp.columns)
       for i in range(0,len(list_GUID)):
           GUID = list_GUID[i]
           temp = df_temp[df_temp[col_guid]==GUID]
           temp = temp[-temp.duplicated()]
           df_guid_subjID = pd.concat([df_guid_subjID, df_temp[df_temp[col_guid]==GUID]],ignore_index = False)
    else:
       # find SubjIDs with multiple GUIDs
       cnt_subjID = df_temp.groupby(col_subjID).nunique()
       list_subjID = cnt_subjID.index[(cnt_subjID[col_guid] > 1) | (cnt_subjID[col_assocguid] > 1)].tolist()
       df_subjID_guid = pd.DataFrame(columns = df_temp.columns)
       for i in range(0,len(list_subjID)):
           subjID = list_subjID[i]
           temp = df_temp[df_temp[col_subjID]==subjID]
           temp = temp[-temp.duplicated(subset = df_temp.drop('record', axis = 1).columns)]
           df_subjID_guid = pd.concat([df_subjID_guid, temp], ignore_index = False)
           
       # find GUIDs with multiple SubjIDs  
       cnt_GUID = df_temp.groupby(col_guid).nunique()
       list_GUID = cnt_GUID.index[(cnt_GUID[col_subjID] > 1) | (cnt_GUID[col_assocguid] > 1)].tolist()
       df_guid_subjID = pd.DataFrame(columns = df_temp.columns)
       for i in range(0,len(list_GUID)):
           GUID = list_GUID[i]
           temp = df_temp[df_temp[col_guid]==GUID]
           temp = temp[-temp.duplicated(subset = df_temp.drop('record', axis = 1).columns)]
           df_guid_subjID = pd.concat([df_guid_subjID, df_temp[df_temp[col_guid]==GUID]],ignore_index = False)
       
    return(df_guid_subjID, df_subjID_guid)

    
####### OUTPUT EXCEL FILE FORMATTING #######

# make excel cols the right width
def get_col_widths(dataframe):
    # First we find the maximum length of the index column   
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max + max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]


# worksheet highlighting/mod class
class Closeout_Highlighting:
    
    # highlighting fill colors
    gray = 'BABABA'
    yellow = 'FFFB10'
    orange = 'FF963A'
    blue = '4BB6FF'
    red = 'FF4B4B'
    gray_fill = PatternFill(fgColor = 'BABABA', fill_type = "solid") # Empty Columns
    yellow_fill = PatternFill(fgColor = 'FFFB10', fill_type = "solid") # Partial Data Elements
    orange_fill = PatternFill(fgColor = 'FF963A', fill_type = "solid") # GUID-SubjID issues
    blue_fill = PatternFill(fgColor = '4BB6FF', fill_type = "solid") # Duplicate Rows
    red_fill = PatternFill(fgColor ='FF4B4B', fill_type = "solid") # Extra-validation Errors
    
    # init method or constructor 
    def __init__(self,  **kwargs):
        self.kwargs = kwargs
        if 'worksheet' in kwargs:
            self.n_cols_excel = self.kwargs['worksheet'].max_column       
            self.n_rows_excel = self.kwargs['worksheet'].max_row
    
    
     # duplicates (BLUE FILL) 
    def duplicates_fill(self, worksheet):
        if not self.kwargs['dup_rows'].empty:
            for row in self.kwargs['dup_rows']:
                    excel_row_idx = row + 1  + 1 # excel row index starts at 1 & header is additional so offset by 2
                    for cols in worksheet.iter_cols(min_col= 1, max_col= self.n_cols_excel, min_row= excel_row_idx, max_row = excel_row_idx):
                        for cell in cols:
                            cell.fill = self.blue_fill
        return worksheet
    
    # partial Data (YELLOW FILL)
    def partialData_fill(self, worksheet):
        if len(self.kwargs['cols_partial_DE']) > 0:
            cols_partial_DE_rg = list(set(self.kwargs['cols_partial_DE']) & set(self.kwargs['cols_rg']))
            cols_partial_DE_nrg = list(set(self.kwargs['cols_partial_DE']) & set(self.kwargs['cols_nrg']))
            cols_excel_partial_DE_rg_idx = [self.kwargs['df_unflat'].columns.get_loc(c) + 1 for c in cols_partial_DE_rg if c in self.kwargs['df_unflat']] # excel col index starts at 1 so offset by 1
            cols_excel_partial_DE_nrg_idx = [self.kwargs['df_unflat'].columns.get_loc(c) + 1 for c in cols_partial_DE_nrg if c in self.kwargs['df_unflat']] # excel col index starts at 1 so offset by 1
            excel_nrg_row_idx = self.kwargs['df_unflat'][self.kwargs['df_unflat']['Repeatable Group'] == 1].index + 1 + 1 # account for excel starting at 1 and header
            for col_idx in cols_excel_partial_DE_nrg_idx: # nrg partial data
                for row_idx in excel_nrg_row_idx:
                    cell = worksheet.cell(row = row_idx, column = col_idx)
                    if cell.value == '': 
                        cell.fill = self.yellow_fill
            for col_idx in cols_excel_partial_DE_rg_idx: # rg partial data
                    for rows in worksheet.iter_rows(min_row=1, max_row= self.n_rows_excel, min_col= col_idx, max_col = col_idx):
                        for cell in rows:
                            if cell.value == '': 
                                cell.fill = self.yellow_fill 
        return worksheet
    
    # subjID-GUID (ORANGE FILL)
    def subjID_guid_fill(self, worksheet):
        subjID_guid_recs = self.kwargs['df_subjID_guid']['record'].tolist()
        subjID_guid_row_idx = self.kwargs['df_unflat'][self.kwargs['df_unflat']['record'].isin(subjID_guid_recs)].index
        if len(subjID_guid_row_idx) > 0:
            col_subjID_idx = self.kwargs['df_unflat'].columns.get_loc(self.kwargs['col_subjID'])
            excel_col_letter = get_column_letter(col_subjID_idx + 1)
            for row in subjID_guid_row_idx:
                 excel_row_idx = row + 1  + 1 # excel row index starts at 1 & header is additional so offset by 2
                 excel_coord = str(excel_col_letter) + str(excel_row_idx)
                 cell = worksheet[excel_coord]
                 cell.fill = self.orange_fill
        return worksheet
                        
    # GUID-subjID (ORANGE FILL)
    def guid_subjID_fill(self, worksheet):
        guid_subjID_recs = self.kwargs['df_guid_subjID']['record'].tolist()
        guid_subjID_row_idx = self.kwargs['df_unflat'][self.kwargs['df_unflat']['record'].isin(guid_subjID_recs)].index
        if len(guid_subjID_row_idx) > 0:
            col_guid_idx = self.kwargs['df_unflat'].columns.get_loc(self.kwargs['col_guid'])
            excel_col_letter = get_column_letter(col_guid_idx + 1)
            for row in guid_subjID_row_idx:
                 excel_row_idx = row + 1  + 1 # excel row index starts at 1 & header is additional so offset by 2
                 excel_coord = str(excel_col_letter) + str(excel_row_idx)
                 cell = worksheet[excel_coord]
                 cell.fill = self.orange_fill
        return worksheet
             
    # emptyColumns (GREY FILL)
    def emptyCols_hide(self, worksheet):
        excel_emptyCol_idx = [self.kwargs['df_unflat'].columns.get_loc(c) + 1 for c in self.kwargs['emptyColumns'] if c in self.kwargs['df_unflat'].columns] # excel col index starts at 1 so offset by 1
        for col_idx in excel_emptyCol_idx:
                for rows in worksheet.iter_rows(min_row=1, max_row= self.n_rows_excel, min_col= col_idx, max_col = col_idx):
                    for cell in rows:
                        cell.fill = self.gray_fill
                worksheet.column_dimensions[get_column_letter(col_idx)].hidden= True
        return worksheet
    
    # validation errors (RED FILL)
    def validationErrors_fill(self, worksheet):
        if self.kwargs['df_val_filt'].empty == False:
            df_errors = self.kwargs['df_val_filt'][self.kwargs['df_val_filt']['Result'] == 'ERROR']
            for idx in range(0, df_errors.shape[0]):
                try:
                    row_temp = int(df_errors.loc[idx,'Validation Row'])-1  # -1 in row to account for first 2 rows in form structure 
                except:
                    row_temp = np.nan
                col_temp = df_errors.loc[idx,'All Submitted Data Column']
                if pd.notnull(row_temp*col_temp): # if null then skip highlighting a coordinate because there is missing row or col info
                    cell = worksheet.cell(row = row_temp, column = col_temp)
                    cell.fill = self.red_fill       
            return worksheet
    
