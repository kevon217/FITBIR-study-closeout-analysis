# STUDY CLOSEOUT ANALYSIS SCRIPT 

# The purpose of this document is help with analyzing the data using the Python Script. The main goals of this analysis will accomplish the following:
# - Identify the number of GUIDS in the dataset
# - Identify duplicate GUIDs appearing in the dataset
# - Identify duplicate data within the dataset
# - Identify missing data within a dataset
# - Check if GUIDs exist in base file for comparison
# - Report validation errors from resultDetails.txt
# - Check if GUIDs and SubjIDs are one-to-one

# Output from this script:
# - Closeout_Analysis Folder:
#    - Study ID_output_excel files:
#        - Separate excel files for each form with analysis results
#    - A summary table of the analysis
#    - FITBIR Study Closeout Report_Template
#    - Log file

# When you run this script you will be prompted to do the following:
# 1- Select starting directory (where this script and associated scripts live).
# 2- Choose whether you will be analyzing flattened QT ("Yes") or original DR files ("No"), then select appropriate directory.
# 3- Answer whether you have validated the reformatted unflattened files in the submission tool? If so ("Yes"), choose resultDetails.txt file
# 4- If you want to filter by DatasetID ("Yes"), choose csv file with list of DatasetIDs.
# 5- Enter Study ID.
# 6- Enter Study name.
# 7- Select BaseFile.csv

# Load packages
import pandas as pd
import datetime 
import sys
import glob
import os
import tkinter as tk # for dialog boxes
from tkinter import messagebox
from tkinter import filedialog
import importlib
from loguru import logger # for logging script actions
logger.remove()
from openpyxl.utils import get_column_letter, column_index_from_string
import re

# Choose starting directory where closeout script files live
root = tk.Tk()
root.withdraw()
start_dir = filedialog.askdirectory(title = 'Select Starting directory where script lives')
os.chdir(start_dir)

# Load Analyze_Validation_Error_Log & fitbir_analysis_functions 
import Analyze_Validation_Error_Log as Analyze_Validation_Error_Log # for analyzing resultDetails.txt file from validation tool
importlib.reload(Analyze_Validation_Error_Log)
import fitbir_analysis_functions as faf # contains custom functions for loading/processing FITBIR csv datafiles 
importlib.reload(faf)

# Query Tool or Data Repostiory Files?
qt_dr_1_0 = messagebox.askyesno("Python","Are your files from Query Tool (yes) or data repository (no)?")
n_skiprows = 0 if qt_dr_1_0 else 1 # skip 1 row if a Data Repository file to remove FS name in A1
if qt_dr_1_0:
    dir_flat = filedialog.askdirectory(title = 'Select Flattened files directory')  # select directory with flattened Query Tool files
    os.chdir(dir_flat)
    dir_parent = os.path.dirname(dir_flat)
else: 
    DR_dir = filedialog.askdirectory(title = 'Select Data Repository files directory') # select directory with unflattened Data Repository files
    os.chdir(DR_dir)
    dir_parent = os.path.abspath(DR_dir)

# Create Closeout_Analysis directory
closeout_analysis_dir = faf.create_folder(dir_parent + '/Closeout_Analysis')

# Create logger file in Closeout_Analysis folder
sep = "," # for separating log messages with comma for opening as csv and filtering
logger_filepath = closeout_analysis_dir + "/Closeout_Log_{time}.log" # need to figure out how to get rid of milliseconds in log file name 
logger.add(logger_filepath, format = "{time:YYYY-MM-DD HH:mm:ss},{level},{message}", backtrace = False, diagnose = False)
logger.add(sys.stderr, colorize = True, format = "<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level: <8}</level> <level>{message}</level>")
logger.info(sep.join(['Closeout scripts folder path:', start_dir]))
logger.info(sep.join(['Closeout_Analysis folder path:', dir_parent]))
    
# Run Analyze_Validation_Error_Log if applicable
val_yes_no = messagebox.askyesno("Python","Do you have unflattened files that were validated in submission tool?")
if val_yes_no == True:
    resultDetails_file = filedialog.askopenfilename(title = 'Select resultDetails.txt file in Validation_CSVs directory', initialdir = os.path.dirname(dir_parent))
    dir_unflat = os.path.dirname(os.path.dirname(os.path.dirname((resultDetails_file))))
    logger.info(sep.join(['Unflattened files directory:', dir_unflat]))
    logger.info(sep.join(['resultDetails.txt validation file:', resultDetails_file]))
    dict_fn_index, dict_FS_errors, df_val = Analyze_Validation_Error_Log.main(resultDetails_file)
else:
    dict_FS_errors = {}
    logger.info(sep.join(['Unflattened files directory:', 'NA']))
    logger.info(sep.join(['resultDetails.txt validation file:', 'NA']))    

# Set working directory for analysis    
if qt_dr_1_0:
    logger.info(sep.join(['Flattened files directory:', dir_flat]))
    logger.info(sep.join(['Data Repository files directory:', 'NA']))
    os.chdir(dir_flat)
else: 
    logger.info(sep.join(['Flattened files directory:', 'NA']))
    logger.info(sep.join(['Data Repository files directory:', DR_dir]))
    os.chdir(DR_dir)    

# Filter by Dataset ID? 
filter_ID_1_0 = messagebox.askyesno("Python","Do you want to filter by Dataset IDs")
if filter_ID_1_0:
    if qt_dr_1_0:
        filter_ID_file = filedialog.askopenfilename(title= 'Choose file with list of Dataset IDs for filter',initialdir = os.path.dirname(dir_flat))  
        logger.info(sep.join(['DatasetID Filter file:', filter_ID_file]))
        filter_ID_list = pd.read_csv(filter_ID_file,low_memory=False).iloc[:,0].dropna().unique().tolist()
    else:
        filter_ID_list = []
        logger.warning(sep.join(['DatasetID Filter file:', 'NA', 'If you are analyzing data repository files each will be its own unique dataset so no need to filter by DatasetID.']))
else:
    logger.info(sep.join(['DatasetID Filter file:', 'NA']))
    filter_ID_list = []

# Notebook settings
pd.options.display.max_seq_items = 2000
pd.options.display.max_rows = 2000
pd.options.display.max_columns = 999    
        
# Set study Id and study name
print('Please enter studyID: ')
studyID=input()
logger.info(sep.join(['Study ID:', studyID]))
print('Please enter study_name: ')
study_name = input()
logger.info(sep.join(['Study name:', study_name]))

# Get filenames of forms and number
filenames = sorted(glob.glob('*.csv'))
num_forms = len(filenames)
logger.info(sep.join(['Number of forms analyzing:', str(num_forms)]))

# Create data frame for summary_table
headers = ["Count","Form Structure","Unique GUIDs",
           "Total # of Records",
           "Total # Standard Form Validation Errors",
           "Total # GUIDs With Partial Data",
           "Duplicate Records", 
           "GUIDs in BaseFile Missing in FS", "GUIDs in FS Missing in BaseFile",
           "GUIDs Mult SubjIDs",
           "SubjIDs Mult GUIDs",
           "FITBIR Comments"]
summary_table = pd.DataFrame(index=range(num_forms),columns=headers)
summary_table["Count"] = range(1,num_forms+1) # add form structure Count

# Choose basefile for GUID demographics comparison across forms
if qt_dr_1_0:
    fn_base_file= filedialog.askopenfilename(title= 'Choose BaseFile for GUID Comparison',initialdir = dir_flat)
    fn_base_file= fn_base_file.split('/')[-1]
    FS_base_file = fn_base_file.split('query_result_')[1].split('_20')[0]
    df_base_file = faf.loadFile(fn_base_file,n_skiprows, filter_ID_1_0, filter_ID_list)[0]
    logger.info(sep.join(['Basefile for GUID comparison:', fn_base_file]))
    if df_base_file.shape[0] == 0:    
        logger.error(sep.join(['Basefile for GUID comparison:','NA', 'Basefile is empty after filtering by dataset ID. Choose a different basefile']))
        sys.exit()
else: 
    fn_base_file= filedialog.askopenfilename(title= 'Choose BaseFile for GUID Comparison',initialdir = DR_dir)
    fn_base_file= fn_base_file.split('/')[-1]
    FS_base_file = fn_base_file
    df_base_file =  pd.read_csv(fn_base_file,low_memory=False,skiprows=n_skiprows)
    logger.info(sep.join(['Basefile for GUID comparison:', fn_base_file]))
    
# Master GUID/SUBJID list for comparing across forms after analysis    
if qt_dr_1_0:
    df_GUIDSubjID_master = pd.DataFrame(columns = ['Form Structure','Study ID','Dataset','GUID','ASSOCIATED GUID','Subject ID'])
else:
    df_GUIDSubjID_master = pd.DataFrame(columns = ['Form Structure','GUID','Subject ID'])

# Create new directory for output
dir_output = os.path.join(os.getcwd(),str(studyID)+"_output_excel_files")
dir_output = faf.create_folder(dir_output)
logger.info(sep.join(['Output excel files path:', dir_output]))

# Start time and Python Version for analysis
py_ver = sep.join(['Python version:', sys.version.replace(",",";")])
pd_ver = sep.join(['Pandas version:', pd.__version__])
logger.info(py_ver)
logger.info(pd_ver)
logger.info('STARTING ANALYSIS...')
start_time = datetime.datetime.now()

# Closeout analysis 
idx = 0
for filename in filenames:
    
    ### READ IN FILE ###
    if qt_dr_1_0:
        formName = filename.split('query_result_')[1].split('_20')[0]
    else:     
        formName = filename
    newFile = "StudyID_" +str(studyID)+ '_' + filename + '_results.xlsx'
    logger.info(sep.join(["LOADING...", filename]))
    df_flat, cols_rg, cols_nrg, col_guid, col_assocguid, col_subjID, n_records = faf.loadFile(filename,n_skiprows, filter_ID_1_0, filter_ID_list)
    if filter_ID_1_0 == 1 and df_flat.shape[0] == 0: # next loop if filtering dataset results in empty df
           logger.warning(sep.join(['File is empty after filtering by dataset ID. No output excel file created:', formName]))
           idx = idx + 1
           continue
    df_unflat = faf.flat2unflat(df_flat, filename, cols_rg, cols_nrg, col_guid, col_subjID, col_assocguid) # if is already flat, will just add 'record' and 'Repeatable Group' columns
    
    ### RUN EMPTY COL/DUPLICATES/PARTIAL DATA ANALYSES ###
    logger.info(sep.join(["SEARCHING FOR EMPTY COLUMNS...", filename]))
    dfReduced, emptyColumns, n_DE_total, n_DE_data, n_GUIDs_uniq = faf.currentDataSubmitted(df_unflat, col_guid)
    logger.info(sep.join(["SEARCHING FOR DUPLICATES...", filename]))
    dfDuplicated, dup_rows, n_dup_rows = faf.returnDuplicated(df_unflat, col_guid)
    logger.info(sep.join(["SEARCHING FOR PARTIAL DATA...", filename]))
    dfMissing,  cols_partial_DE, guidCounts = faf.GUIDswMissingDatainDEs(df_flat, cols_rg, cols_nrg) # use flattened file since rg will have lot of empty elements
    
    ### VALIDATION RESULTS MODIFIED FOR 'ALL SUBMITTED DATA' TAB HIGHLIGHTING ###
    if val_yes_no == True:
        if formName in dict_fn_index.keys():
            df_val_filt = df_val.iloc[dict_fn_index.get(formName)[0]:dict_fn_index.get(formName)[1],:].copy()
            df_val_filt = df_val_filt[((df_val_filt['Result'] == 'ERROR') | (df_val_filt['Result'] == 'WARNING')) &  
                                      ((df_val_filt['GUID'].isnull() != True) | (df_val_filt['Validation Row'].isnull() != True))].copy().reset_index(drop=True) # reduce df to rows with ERRORS or WARNINGS
            df_val_filt = df_val_filt.drop(columns = ['FS File'])
            if qt_dr_1_0:
                if len(col_assocguid) == 0:
                    df_val_filt['All Submitted Data Column'] = df_val_filt['All Submitted Data Column'].apply(lambda x: x if pd.isna(x) else column_index_from_string(x)+4) # adjust for column offset of Record/Repeatable Group/Validation Row/StudyID + excel index starts at 1
                else:     
                    df_val_filt['All Submitted Data Column'] = df_val_filt['All Submitted Data Column'].apply(lambda x: x if pd.isna(x) else column_index_from_string(x)+5) # adjust for column offset of Record/Repeatable Group/Validation Row/StudyID + excel index starts at 1
                df_val_filt['Validation Row'] = df_val_filt['Validation Row'].apply(lambda x: x if pd.isna(x) else int(x)) # preserve row number from validation report
                logger.info(sep.join(["EXTRACTING EXTRA-VALIDATION RESULTS...", filename]))
            else:
                df_val_filt['All Submitted Data Column'] = df_val_filt['All Submitted Data Column'].apply(lambda x: x if pd.isna(x) else column_index_from_string(x)+2) # adjust for DR files
        else: 
            df_val_filt = pd.DataFrame()
            logger.info(sep.join(["NO EXTRA-VALIDATION PERFORMED...", filename]))    
    else:
        df_val_filt = pd.DataFrame()
        logger.info(sep.join(["NO EXTRA-VALIDATION PERFORMED...", filename]))    

    ### START FILLING IN PART OF SUMMARY TABLE ###
    summary_table.at[idx,"Form Structure"] = formName
    summary_table.at[idx,"Unique GUIDs"] = n_GUIDs_uniq
    summary_table.at[idx,"Total # of Records"] = n_records
    if formName in dict_FS_errors.keys():
        summary_table.at[idx,"Total # Standard Form Validation Errors"] = dict_FS_errors.get(formName) # if not in dictionary, will report "None"
    else:    
        summary_table.at[idx,"Total # Standard Form Validation Errors"] = "No extra-validation."
    summary_table.at[idx,"Total # GUIDs With Partial Data"] = len(dfMissing) # this may not be necessary what about number of DEs with partial
    summary_table.at[idx,"Duplicate Records"] = n_dup_rows # /2 # account for double counting 
    
    ### GUID DEMOGRAPHICS CHECK ###
    logger.info(sep.join(["SEARCHING FOR MISSING GUIDS...", filename]))
    if formName == FS_base_file:
        summary_table.loc[(summary_table['Form Structure'] == FS_base_file),"GUIDs in BaseFile Missing in FS"] = "(BASE FILE)"
        summary_table.loc[(summary_table['Form Structure'] == FS_base_file),"GUIDs in FS Missing in BaseFile"] = "(BASE FILE)"
    else: 
        # write GUID comparison
        df_GUIDs_miss_compfile, n_GUIDs_miss_compfile = faf.GUIDsMissingFromCompFile(fn_base_file,filename,df_base_file,df_flat)
        df_GUIDs_miss_basefile, n_GUIDs_miss_basefile = faf.GUIDsMissingFromBase(fn_base_file,filename,df_base_file,df_flat)
        summary_table.loc[(summary_table['Form Structure'] == formName),"GUIDs in BaseFile Missing in FS"] = n_GUIDs_miss_compfile       
        summary_table.loc[(summary_table['Form Structure'] == formName),"GUIDs in FS Missing in BaseFile"] = n_GUIDs_miss_basefile
    
    ### GUID - SUBJID MATCH CHECK ###
    logger.info(sep.join(["COMPARING GUIDS AND SUBJECT IDS... ", filename]))
    try: # temporary work around for form structures that don't have subjectIDs e.g., MDS_UPDRS
        df_guid_subjID, df_subjID_guid = faf.guid_subjID_check(df_flat, qt_dr_1_0, col_guid, col_assocguid, col_subjID)
        n_guids_Mult_subjids =  len(df_guid_subjID.filter(regex = re.compile("GUID", re.IGNORECASE)).iloc[:,0].unique())
        n_subjids_Mult_guids =  len(df_subjID_guid.filter(regex = re.compile("Subject[_ ]?ID", re.IGNORECASE)).iloc[:,0].unique())
        summary_table.at[idx,"GUIDs Mult SubjIDs"] = n_guids_Mult_subjids
        summary_table.at[idx,"SubjIDs Mult GUIDs"] = n_subjids_Mult_guids
        if qt_dr_1_0:
            if len(col_assocguid) == 0:
                df_GUIDSubjID_temp = df_flat[['Study ID','Dataset', col_guid, col_subjID]].copy()
                df_GUIDSubjID_temp.rename(columns= {col_guid:'GUID',col_subjID:'Subject ID'}, inplace = True)
            else:
                df_GUIDSubjID_temp = df_flat[['Study ID','Dataset', col_guid, col_assocguid, col_subjID]].copy()
                df_GUIDSubjID_temp.rename(columns= {col_guid:'GUID', col_assocguid:'ASSOCIATED GUID',col_subjID:'Subject ID'}, inplace = True)
        else:
            df_GUIDSubjID_temp = df_flat[[col_guid, col_subjID]].copy()
            df_GUIDSubjID_temp.rename(columns= {col_guid:'GUID',col_subjID:'Subject ID'}, inplace = True)
        df_GUIDSubjID_temp.insert(0,'Form Structure', formName)
        df_GUIDSubjID_master = pd.concat([df_GUIDSubjID_master, df_GUIDSubjID_temp])
    except: 
        df_subjID_guid = []
        df_guid_subjID = []
        logger.warning(sep.join(["Form structure may not have a SubjectIDs data element:", formName]))
    
    ### CREATE OUTPUT EXCEL FILE ###
    logger.info("WRITING OUTPUT EXCEL FILE AND SHEETS... ")
    writer = pd.ExcelWriter(dir_output+'\\'+newFile, engine = 'openpyxl')
    max_col_excel = 16384 
    max_row_excel = 1048576
    ws_highlight_obj = faf.Closeout_Highlighting()
    
    # write all data sheet / highlight all findings
    sheet_name = 'All Submitted Data'
    if df_unflat.shape[0] <= max_row_excel and df_unflat.shape[1] <= max_col_excel:
        df_unflat.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        for i, width in enumerate(faf.get_col_widths(df_unflat)):
            excel_col_letter = get_column_letter(i+1)
            worksheet.column_dimensions[excel_col_letter].width = width
        ws_temp = faf.Closeout_Highlighting(worksheet = worksheet, formName = formName, df_unflat = df_unflat, df_val_filt = df_val_filt,
                                            emptyColumns = emptyColumns, cols_partial_DE = cols_partial_DE, cols_rg = cols_rg, 
                                            cols_nrg = cols_nrg, dup_rows = dup_rows, df_guid_subjID = df_guid_subjID, df_subjID_guid = df_subjID_guid, 
                                            col_guid = col_guid, col_subjID = col_subjID)
        worksheet = ws_temp.duplicates_fill(worksheet)
        worksheet = ws_temp.partialData_fill(worksheet)
        if isinstance( df_subjID_guid, pd.DataFrame) and isinstance(df_guid_subjID, pd.DataFrame):
            worksheet = ws_temp.subjID_guid_fill(worksheet) 
            worksheet = ws_temp.guid_subjID_fill(worksheet) 
        worksheet = ws_temp.emptyCols_hide(worksheet) 
        worksheet = ws_temp.validationErrors_fill(worksheet) 
    else:
        logger.warning(sep.join(["Sheet too large to write to excel:", sheet_name, "Shape = " + str(df_unflat.shape).replace(",",";")]))

    # write parsed/adjusted validation report sheet
    if val_yes_no == True:
        if formName in dict_fn_index.keys():
            sheet_name = 'Validation Report'
            if df_val_filt.shape[0] <= max_row_excel and df_val_filt.shape[1] <= max_col_excel:
                df_val_filt.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                worksheet.sheet_properties.tabColor = ws_highlight_obj.red
                for i, width in enumerate(faf.get_col_widths(df_unflat)):
                   excel_col_letter = get_column_letter(i+1)
                   worksheet.column_dimensions[excel_col_letter].width = width
            else:
                logger.warning(sep.join(["Sheet too large to write to excel:", sheet_name, "Shape = " + str(df_unflat.shape).replace(",",";")]))
                
    # write empty columns sheet
    sheet_name = 'Empty Data Elements'
    try:
        df_emptyColumns = pd.DataFrame(emptyColumns).rename(columns={0:'Empty Data Elements'})
        df_emptyColumns.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        worksheet.sheet_properties.tabColor = ws_highlight_obj.gray
        for i, width in enumerate(faf.get_col_widths(df_emptyColumns)):
            excel_col_letter = get_column_letter(i+1)
            worksheet.column_dimensions[excel_col_letter].width = width
    except:
        logger.warning(sep.join(["Sheet too large to write to excel:", sheet_name, "Length = " + str(len(emptyColumns))]))

    # write missing data sheet
    sheet_name = 'DEs with Partial Data'
    try:
        dfMissing.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        worksheet.sheet_properties.tabColor = ws_highlight_obj.yellow
        for i, width in enumerate(faf.get_col_widths(dfMissing)):
            excel_col_letter = get_column_letter(i+1)
            worksheet.column_dimensions[excel_col_letter].width = width
    except: 
        logger.warning(sep.join(["Sheet too large to write to excel:", sheet_name, "Shape = " + str(dfMissing.shape).replace(",",";")]))

    # write duplicated sheet
    sheet_name = 'Duplicated Data'
    try: 
        dfDuplicated.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        worksheet.sheet_properties.tabColor = ws_highlight_obj.blue
        for i, width in enumerate(faf.get_col_widths(dfDuplicated)):
            excel_col_letter = get_column_letter(i+1)
            worksheet.column_dimensions[excel_col_letter].width = width
        ws_temp = faf.Closeout_Highlighting(worksheet = worksheet, formName = formName, df_unflat = df_unflat, df_val_filt = df_val_filt,
                                            emptyColumns = emptyColumns, cols_partial_DE = cols_partial_DE, cols_rg = cols_rg, 
                                            cols_nrg = cols_nrg, dup_rows = dup_rows, df_guid_subjID = df_guid_subjID, df_subjID_guid = df_subjID_guid, 
                                            col_guid = col_guid, col_subjID = col_subjID)
        worksheet = ws_temp.emptyCols_hide(worksheet)
        
    except:
        logger.warning(sep.join(["Sheet too large to write to excel:", sheet_name, "Shape = " + str(dfDuplicated.shape).replace(",",";")]))

    # write GUID counts data sheet
    sheet_name = 'GUID Counts in this Form'
    try:
        guidCounts = pd.DataFrame(guidCounts).reset_index()
        guidCounts.columns.values[0] = 'GUID'
        guidCounts.columns.values[1] = 'GUID counts'    
        guidCounts.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        for i, width in enumerate(faf.get_col_widths(guidCounts)):
            excel_col_letter = get_column_letter(i+1)
            worksheet.column_dimensions[excel_col_letter].width = width
    except:
        logger.warning(sep.join(["Sheet too large to write to excel:", sheet_name, "Length = " + str(len(guidCounts))]))

     # write BASEFILE vs. FS GUID comparisons
    if formName == FS_base_file:
        logger.warning(sep.join(["No need to compare Basefile with itself. Excluding GUID comparison sheets.", FS_base_file]))
    else: 
        sheet_name = FS_base_file +' not in '+formName
        sheet_name = (sheet_name[:28]+'...') if len(sheet_name) > 31 else sheet_name
        try: 
            df_GUIDs_miss_compfile.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for i, width in enumerate(faf.get_col_widths(df_GUIDs_miss_compfile)):
                excel_col_letter = get_column_letter(i+1)
                worksheet.column_dimensions[excel_col_letter].width = width
        except: 
            logger.warning(sep.join(["Sheet too large to write to excel:", sheet_name, "Shape = " + str(df_GUIDs_miss_compfile.shape).replace(",",";")]))
    
        sheet_name = formName +' not in '+ FS_base_file
        sheet_name = (sheet_name[:28]+'...') if len(sheet_name) > 31 else sheet_name
        try:
            df_GUIDs_miss_basefile.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for i, width in enumerate(faf.get_col_widths(df_GUIDs_miss_basefile)):
                excel_col_letter = get_column_letter(i+1)
                worksheet.column_dimensions[excel_col_letter].width = width
        except: 
            logger.warning(sep.join(["Sheet too large to write to excel:", sheet_name, "Shape = " + str(df_GUIDs_miss_basefile.shape).replace(",",";")]))
        
    # write same subjID -> different GUIDS or same GUID different subjID
    sheet_name = 'GUIDs Mult SubjIDs'
    try: 
        df_guid_subjID.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        worksheet.sheet_properties.tabColor = ws_highlight_obj.orange
        for i, width in enumerate(faf.get_col_widths(df_guid_subjID)):
            excel_col_letter = get_column_letter(i+1)
            worksheet.column_dimensions[excel_col_letter].width = width
    except: 
        try:
            logger.info(sep.join(["Sheet too large to write to excel:", sheet_name, "Shape = " + str(df_guid_subjID.shape).replace(",",";")]))
        except:
            logger.warning(sep.join(["SubjectIDs don't exist in:", formName]))
    sheet_name = 'SubjIDs Mult GUIDs'
    try:
        df_subjID_guid.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        worksheet.sheet_properties.tabColor = ws_highlight_obj.orange
        for i, width in enumerate(faf.get_col_widths(df_subjID_guid)):
            excel_col_letter = get_column_letter(i+1)
            worksheet.column_dimensions[excel_col_letter].width = width
    except: 
        try:
            logger.info(sep.join(["Sheet too large to write to excel:", sheet_name, "Shape = " + str(df_subjID_guid.shape).replace(",",";")]))
        except:
            logger.warning(sep.join(["SubjectIDs don't exist in:", formName]))
            
    # write excel file if non-empty dataframe
    if df_unflat.empty != True:
        # writer.save() # this was causing repair notification issue
        writer.close()
    else:
        logger.warning(sep.join(['File is empty. No output excel file created:', formName]))
        
    idx = idx + 1
      

# Review GUID-SubjID across forms
df_guid_subjID, df_subjID_guid = faf.guid_subjID_check(df_GUIDSubjID_master, qt_dr_1_0, 'GUID', 'ASSOCIATED GUID', 'Subject ID')
if qt_dr_1_0: 
    df_guid_subjID = df_guid_subjID.drop_duplicates(['Dataset','GUID','ASSOCIATED GUID','Subject ID']).drop('record', axis = 1).sort_values('GUID')
    df_subjID_guid = df_subjID_guid.drop_duplicates(['Dataset','GUID','ASSOCIATED GUID','Subject ID']).drop('record', axis = 1).sort_values('Subject ID')
else:
    df_guid_subjID = df_guid_subjID.drop_duplicates(['Form Structure','GUID','Subject ID']).drop('record', axis = 1).sort_values('GUID')
    df_subjID_guid = df_subjID_guid.drop_duplicates(['Form Structure','GUID','Subject ID']).drop('record', axis = 1).sort_values('Subject ID')


# Create excel summary table file
logger.info('WRITING SUMMARY TABLE EXCEL FILE...')
summary_table_fn = "StudyID_" + studyID + "_Closeout_Analysis_Summary_Table.xlsx"
writer = pd.ExcelWriter(os.path.join(closeout_analysis_dir, summary_table_fn), engine = 'openpyxl')

# Write summary table
sheet_name = "Summary Table"
summary_table.to_excel(writer, sheet_name= sheet_name, index=False)
worksheet = writer.sheets[sheet_name]
for i, width in enumerate(faf.get_col_widths(summary_table)):
    excel_col_letter = get_column_letter(i+1)
    worksheet.column_dimensions[excel_col_letter].width = width

# Write GUID-SubjectID Master List
sheet_name = 'GUID-Subject ID Master List'
df_GUIDSubjID_master.drop_duplicates().to_excel(writer, sheet_name= sheet_name, index=False)
worksheet = writer.sheets[sheet_name]
for i, width in enumerate(faf.get_col_widths(df_guid_subjID)):
    excel_col_letter = get_column_letter(i+1)
    worksheet.column_dimensions[excel_col_letter].width = width

# Write GUIDs with multiple subject IDs across forms
sheet_name = 'GUID multiple Subject IDs'
df_guid_subjID.to_excel(writer, sheet_name= sheet_name, index=False)
worksheet = writer.sheets[sheet_name]
for i, width in enumerate(faf.get_col_widths(df_guid_subjID)):
    excel_col_letter = get_column_letter(i+1)
    worksheet.column_dimensions[excel_col_letter].width = width

# Write subject IDs with multiple GUIDs across forms    
sheet_name = 'Subject ID multiple GUIDs'
df_subjID_guid.to_excel(writer, sheet_name= sheet_name, index=False)
worksheet = writer.sheets[sheet_name]
for i, width in enumerate(faf.get_col_widths(df_subjID_guid)):
    excel_col_letter = get_column_letter(i+1)
    worksheet.column_dimensions[excel_col_letter].width = width

# Close excel file 
writer.close() # removed writer.save() as this was causing repair notification issue

# Create closeout analysis package
logger.info('GATHERING CLOSEOUT FILES...')
dir_output_copy = closeout_analysis_dir + '/' + os.path.basename(dir_output)
faf.copyDirectory(dir_output,dir_output_copy)
template_path = start_dir + '/FITBIR Study Closeout Report_Template.docx'
faf.shutil.copy(template_path,closeout_analysis_dir)

# End of script
logger.info(sep.join(['Script completion time:', str(datetime.datetime.now() - start_time)]))  
logger.remove()
os.chdir(start_dir)



